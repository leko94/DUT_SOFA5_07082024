import os
import dash
from dash import dcc, html
from dash.dependencies import Input, Output, State
import pandas as pd
import plotly.graph_objects as go
import plotly.express as px
from openpyxl import load_workbook
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from sklearn.linear_model import LinearRegression
from sklearn.impute import SimpleImputer
import numpy as np

# Initialize the Dash app
app = dash.Dash(__name__, suppress_callback_exceptions=True)
server = app.server  # Expose the server for WSGI

# File paths to the Excel files
staff_file_path = 'Chart in Microsoft PowerPoint.xlsx'
students_file_path = 'Students.xlsx'
student_performance_file_path = 'Student Perfomances.xlsx'
dut_file_path = 'DUT Research.xlsx'

# Load the Excel file for DUT Research
df_sheet1 = pd.read_excel(dut_file_path, sheet_name='Sheet1')
df_sheet2 = pd.read_excel(dut_file_path, sheet_name='Sheet2')
df_sheet3 = pd.read_excel(dut_file_path, sheet_name='Sheet3')
df_sheet4 = pd.read_excel(dut_file_path, sheet_name='Sheet4')

# Functions to create charts for staff data
def create_staff_charts():
    workbook = load_workbook(staff_file_path, data_only=True)
    sheet_names = workbook.sheetnames

    # Extract data from each sheet
    titles = {}
    x_labels = {}
    y_labels = {}
    dfs = {}

    for sheet_name in sheet_names:
        sheet = workbook[sheet_name]
        titles[sheet_name] = sheet['A1'].value
        df = pd.read_excel(staff_file_path, sheet_name=sheet_name, header=None, skiprows=0)
        x_labels[sheet_name] = df.iloc[1, 0]  # X-axis label from A2
        y_labels[sheet_name] = df.iloc[1, 1]  # Y-axis label from B2
        dfs[sheet_name] = df.iloc[2:, [0, 1]].rename(columns={0: x_labels[sheet_name], 1: y_labels[sheet_name]})

    # Create bar charts
    def create_bar_chart(df, title, x_label, y_label, color):
        fig = px.bar(
            df,
            x=df.columns[0],
            y=df.columns[1],
            title=title,
            labels={df.columns[0]: x_label, df.columns[1]: y_label},
            color_discrete_sequence=[color]
        )
        fig.update_layout(
            autosize=False,
            width=600,
            height=600,
            bargap=0.2
        )
        fig.update_traces(text=df[df.columns[1]], textposition='outside')  # Add values on bars
        return fig

    # Create line charts
    def create_line_chart(df, title, x_label, y_label, color):
        fig = px.line(
            df,
            x=df.columns[0],
            y=df.columns[1],
            title=title,
            labels={df.columns[0]: x_label, df.columns[1]: y_label},
            color_discrete_sequence=[color]
        )
        fig.update_layout(
            autosize=False,
            width=600,
            height=600
        )
        return fig

    # Create forecast charts
    def create_forecast_chart(df, title, x_label, y_label, color):
        df[x_label] = pd.to_numeric(df[x_label], errors='coerce')
        df[y_label] = pd.to_numeric(df[y_label], errors='coerce')

        # Linear Regression for Forecasting
        X = df[[x_label]].values.reshape(-1, 1)
        y = df[y_label].values
        model = LinearRegression().fit(X, y)

        # Forecast for 2030
        future_years = np.array([2030]).reshape(-1, 1)
        forecast_values = model.predict(future_years)

        fig = px.line(
            df,
            x=df.columns[0],
            y=df.columns[1],
            title=title,
            labels={df.columns[0]: x_label, df.columns[1]: y_label},
            color_discrete_sequence=[color]
        )
        forecast_df = pd.DataFrame({x_label: [2030], y_label: forecast_values})
        fig.add_scatter(x=forecast_df[x_label], y=forecast_df[y_label], mode='markers+text', text=['2030'], textposition='top center', marker=dict(color='red', size=10))
        fig.update_layout(
            autosize=False,
            width=600,
            height=600
        )
        return fig

    # Define custom colors for each chart
    colors = ['blue', 'green', 'orange', 'purple']

    # Create figures for each sheet
    bar_figures = {}
    line_figures = {}
    forecast_figures = {}

    for i, sheet_name in enumerate(sheet_names):
        color = colors[i % len(colors)]
        df = dfs[sheet_name]
        title = titles[sheet_name]
        x_label = x_labels[sheet_name]
        y_label = y_labels[sheet_name]

        if sheet_name in ['Sheet1', 'Sheet2', 'Sheet3']:
            bar_figures[sheet_name] = create_bar_chart(df, title, x_label, y_label, color)
            forecast_figures[sheet_name] = create_forecast_chart(df, f"Forecast for 2030 - {title}", x_label, y_label, color)
        elif sheet_name == 'Sheet4':
            line_figures[sheet_name] = create_line_chart(df, title, x_label, y_label, color)

    # Load data for Sheet5
    df_sheet5 = pd.read_excel(staff_file_path, sheet_name='Sheet5')

    # Extract year columns and data for Sheet5
    years = df_sheet5.columns[1:]
    departments = df_sheet5.iloc[:, 0]
    data = df_sheet5.iloc[:, 1:]

    # Convert data to appropriate format for Sheet5
    df_data = data.copy()
    df_data.columns = years
    df_data['Department'] = departments.values
    df_data = df_data.melt(id_vars='Department', var_name='Year', value_name='Percentage')

    # Extract 2014 and 2022 data
    df_2014 = df_data[df_data['Year'] == '2014'][['Department', 'Percentage']].set_index('Department')
    df_2022 = df_data[df_data['Year'] == '2022'][['Department', 'Percentage']].set_index('Department')

    # Calculate the difference between 2022 and 2014
    percentage_diff = df_2022.join(df_2014, lsuffix='_2022', rsuffix='_2014')
    percentage_diff['Percentage Difference'] = percentage_diff['Percentage_2022'] - percentage_diff['Percentage_2014']

    # Reset index and prepare data for the graph
    percentage_diff = percentage_diff.reset_index()
    percentage_diff.columns = ['Department', 'Percentage 2022', 'Percentage 2014', 'Percentage Difference']

    # Create the bar graph showing the percentage difference
    fig_diff = px.bar(
        percentage_diff,
        x='Department',
        y='Percentage Difference',
        title='Percentage Difference (2022 vs. 2014) for Academic Staff with PhD',
        labels={'Percentage Difference': 'Percentage Difference'},
        height=600,
        color='Department'
    )
    fig_diff.update_traces(text=percentage_diff['Percentage Difference'], textposition='outside')
    fig_diff.update_layout(
        legend_title='Departments',
        autosize=False,
        width=800,
        height=600
    )

    # Create the bar graph for Sheet5 without text labels
    fig_sheet5 = px.bar(
        df_data,
        x='Year',
        y='Percentage',
        color='Department',
        title='Percentage of Full-Time Permanent Academic Staff with PhD (2014-2022)',
        labels={'Percentage': 'Percentage'},
        height=600
    )
    fig_sheet5.update_traces(textposition='none')  # Remove text annotations

    return bar_figures, line_figures, forecast_figures, fig_diff, fig_sheet5

# Functions to create charts for students data
def create_students_charts():
    # Load data from Sheet1
    df1 = pd.read_excel(students_file_path, sheet_name='Sheet1')
    df1.columns = ['Year', 'Actual', 'Planned']
    df1['Actual'] = pd.to_numeric(df1['Actual'], errors='coerce')
    df1['Planned'] = pd.to_numeric(df1['Planned'], errors='coerce')
    df1 = df1.dropna()

    # Create the first line chart
    fig1 = px.line(
        df1,
        x='Year',
        y=['Actual', 'Planned'],
        labels={'value': 'Headcount', 'variable': 'Type'},
        title='Headcount Enrolment: Planned vs Achieved (2014-2023)',
        markers=True
    )
    fig1.update_layout(
        autosize=False,
        width=800,
        height=600
    )

    # Create the second chart with linear regression forecast
    def create_linear_regression_forecast_chart(df):
        df['Year'] = pd.to_numeric(df['Year'], errors='coerce')
        df['Actual'] = pd.to_numeric(df['Actual'], errors='coerce')
        df = df.dropna()
        X = df[['Year']].values.reshape(-1, 1)
        y = df['Actual'].values
        model = LinearRegression().fit(X, y)
        future_years = np.array([2030]).reshape(-1, 1)
        forecast_value = model.predict(future_years)[0]
        fig2 = px.line(
            df,
            x='Year',
            y='Actual',
            title='Linear Regression Forecast for 2030',
            labels={'Year': 'Year', 'Actual': 'Headcount'},
            markers=True
        )
        fig2.add_scatter(
            x=[2030],
            y=[forecast_value],
            mode='markers+text',
            text=['2030'],
            textposition='top right',
            marker=dict(color='red', size=10),
            name='Forecast'
        )
        return fig2

    fig2 = create_linear_regression_forecast_chart(df1)

    # Load data from Sheet2 and get the title from cell A1
    df2 = pd.read_excel(students_file_path, sheet_name='Sheet2', header=1)  # Skip the header row
    title2 = pd.read_excel(students_file_path, sheet_name='Sheet2', header=None).iloc[0, 0]  # Get the title from cell A1

    # Compute the Difference
    df2['Difference'] = df2['Actual'] - df2['Planned']

    # Create the bar graph for Sheet2
    fig3 = px.bar(
        df2,
        x='Departments',
        y=['Planned', 'Actual'],
        title=title2,
        labels={'value': 'Number of Students', 'variable': 'Type'},
        text='Difference'  # Display the difference on the bars
    )
    fig3.update_layout(
        autosize=False,
        width=800,
        height=600
    )

    # Load data from Sheet3
    df3 = pd.read_excel(students_file_path, sheet_name='Sheet3', header=None)

    # Extract title and relevant columns
    title3 = df3.iloc[0, 0]  # Title from cell A1

    # Check the number of columns and set the column names accordingly
    if df3.shape[1] == 4:
        df3.columns = ['Department', '2014', '2022', 'Growth']  # Set column names manually
        df3 = df3.iloc[2:]  # Skip rows before actual data
    elif df3.shape[1] == 5:
        df3.columns = ['Department', '2014', '2022', '2023', 'Growth']  # Set column names manually
        df3 = df3.iloc[2:]  # Skip rows before actual data
    else:
        raise ValueError(f"Unexpected number of columns in Sheet3: {df3.shape[1]}")

    # Convert columns to numeric values and round the Growth values
    df3['2014'] = pd.to_numeric(df3['2014'], errors='coerce')
    df3['2022'] = pd.to_numeric(df3['2022'], errors='coerce')
    if '2023' in df3.columns:
        df3['2023'] = pd.to_numeric(df3['2023'], errors='coerce')
    df3['Growth'] = pd.to_numeric(df3['Growth'], errors='coerce').round(2)

    # Melt the DataFrame for better plotting
    if '2023' in df3.columns:
        df3_melted = df3.melt(id_vars='Department', value_vars=['2014', '2022', '2023'], var_name='Year', value_name='Value')
    else:
        df3_melted = df3.melt(id_vars='Department', value_vars=['2014', '2022'], var_name='Year', value_name='Value')

    # Create the bar graph for Sheet3
    fig4 = px.bar(
        df3_melted,
        x='Department',
        y='Value',
        color='Year',
        text=df3_melted['Department'].map(df3.set_index('Department')['Growth']),  # Add rounded Growth as text
        title=title3,
        labels={'Value': 'Number of Students', 'Year': 'Year'},
        color_discrete_map={'2014': 'blue', '2022': 'green', '2023': 'orange'} if '2023' in df3.columns else {'2014': 'blue', '2022': 'green'}
    )
    fig4.update_traces(texttemplate='%{text:.2f}', textposition='outside')
    fig4.update_layout(
        autosize=False,
        width=800,
        height=600,
        legend_title_text='Year'
    )

    # Load data from Sheet4
    df4 = pd.read_excel(students_file_path, sheet_name='Sheet4')

    # Extract title and labels
    title4 = '% African Students'  # Title from cell A1
    x_label4 = 'Year'  # X-axis label from cell A2
    y_label4 = '% African Students'  # Y-axis label from cell B2

    # Set headers and extract data
    df4.columns = [x_label4, y_label4]
    df4 = df4.iloc[1:]  # Skip rows before actual data

    # Convert '% African Students' to float after stripping '%'
    df4[y_label4] = df4[y_label4].astype(str).str.rstrip('%').astype(float)

    # Create the line graph for Sheet4
    fig5 = px.line(
        df4,
        x=x_label4,
        y=y_label4,
        title=title4,
        labels={x_label4: 'Year', y_label4: '% African Students'},
        markers=True
    )
    fig5.update_layout(
        autosize=False,
        width=800,
        height=600
    )

    # Load data from Sheet5
    df5 = pd.read_excel(students_file_path, sheet_name='Sheet5')

    # Extract title and labels
    title5 = "Percentage Female Students"  # Title from cell A1
    x_label5 = 'Year'  # X-axis label from cell A2
    y_label5 = 'Percentage Female Students'  # Y-axis label from cell B2

    # Set headers and extract data
    df5.columns = [x_label5, y_label5]
    df5 = df5.iloc[1:]  # Skip rows before actual data

    # Convert '% Female Students' to float after stripping '%'
    df5[y_label5] = df5[y_label5].astype(str).str.rstrip('%').astype(float)

    # Create the line graph for Sheet5 with a different color
    fig6 = px.line(
        df5,
        x=x_label5,
        y=y_label5,
        title=title5,
        labels={x_label5: 'Year', y_label5: 'Percentage Female Students'},
        markers=True,
        line_shape='linear'
    )
    fig6.update_layout(
        autosize=False,
        width=800,
        height=600,
        plot_bgcolor='lightgray'  # Different color for distinction
    )

    # Load data from Sheet6
    df6 = pd.read_excel(students_file_path, sheet_name='Sheet6')

    # Extract title and labels
    title6 = "Faculty Postgraduate Enrolment"  # Title from cell A1
    x_label6 = 'Year'  # X-axis label from cell A2
    y_label6 = '% Enrolment'  # Y-axis label from cell B2

    # Set headers and extract data
    df6.columns = [x_label6, y_label6]
    df6 = df6.iloc[1:]  # Skip rows before actual data

    # Convert '% Enrolment' to float after stripping '%'
    df6[y_label6] = df6[y_label6].astype(str).str.rstrip('%').astype(float)

    # Create the line graph
    fig7 = go.Figure()

    # Add the line trace
    fig7.add_trace(go.Scatter(
        x=df6[x_label6],
        y=df6[y_label6],
        mode='lines',
        line=dict(color='purple'),  # Set the color to purple
        name='Enrolment Percentage'
    ))

    # Update layout for the graph from Sheet6
    fig7.update_layout(
        title=title6,
        xaxis_title=x_label6,
        yaxis_title=y_label6,
        autosize=False,
        width=800,
        height=600,
        plot_bgcolor='lightgray'  # Different color for distinction
    )

    # Load data from Sheet7
    df7 = pd.read_excel(students_file_path, sheet_name='Sheet7')
    df7.columns = df7.columns.str.strip()

    # Create the bar chart with actual percentages on top of each bar
    fig8 = px.bar(
        df7,
        x='Department',
        y=['UG', 'PG upto Masters', 'PG'],
        title='Enrolment by Level',
        labels={'value': 'Percentage', 'variable': 'Enrolment Level'},
        barmode='group',
        text_auto=True  # Add text_auto=True to display values on bars
    )

    # Load data from Sheet8
    df8 = pd.read_excel(students_file_path, sheet_name='Sheet8')
    df8.columns = df8.columns.str.strip()

    # Melt the DataFrame for better plotting
    df8_melted = df8.melt(id_vars='Departments', value_vars=df8.columns[1:-1], var_name='Year', value_name='Percentage')
    df8_melted['Percentage'] = df8_melted['Percentage'].astype(str).str.rstrip('%').astype(float)

    # Create the line graph for Sheet8
    fig9 = px.line(
        df8_melted,
        x='Year',
        y='Percentage',
        color='Departments',
        title='Postgraduate (M+D) Enrolment',
        markers=True
    )
    fig9.update_layout(
        autosize=False,
        width=800,
        height=600,
        legend_title_text='Departments'
    )

    # Create the bar graph for Departments and Difference 2014 vs 2022
    fig10 = px.bar(
        df8,
        x='Departments',
        y='Difference 2014 vs 2022',
        title='Difference 2014 vs 2022 by Department',
        labels={'Departments': 'Departments', 'Difference 2014 vs 2022': 'Difference'},
        text='Difference 2014 vs 2022'
    )
    fig10.update_layout(
        autosize=False,
        width=800,
        height=600
    )
    fig10.update_traces(texttemplate='%{text:.2f}', textposition='outside')

    # Load data from Sheet9
    df9 = pd.read_excel(students_file_path, sheet_name='Sheet9')
    df9.columns = df9.columns.str.strip()

    # Melt the DataFrame for better plotting
    df9_melted = df9.melt(id_vars='Department', value_vars=df9.columns[1:], var_name='Year', value_name='No. of Postgraduate enrolment')

    # Create the bar graph for Sheet9
    fig11 = px.bar(
        df9_melted,
        x='Year',
        y='No. of Postgraduate enrolment',
        color='Department',
        title='Postgraduate Enrolment - Actual Student Numbers',
        text='No. of Postgraduate enrolment',
        barmode='group'
    )
    fig11.update_layout(
        autosize=False,
        width=1000,  # Increased the width
        height=700,  # Increased the height
        legend_title_text='Departments'
    )
    fig11.update_traces(texttemplate='%{text}', textposition='outside')

    # Load data from Sheet10
    df10 = pd.read_excel(students_file_path, sheet_name='Sheet10')
    df10.columns = df10.columns.str.strip()

    # Melt the DataFrame for better plotting
    df10_melted = df10.melt(id_vars='Department', value_vars=df10.columns[1:], var_name='Year', value_name='Percentage')
    df10_melted['Percentage'] = df10_melted['Percentage'].astype(str).str.rstrip('%').astype(float)

    # Create the bar graph for Sheet10
    fig12 = px.bar(
        df10_melted,
        x='Year',
        y='Percentage',
        color='Department',
        title='International student Postgraduate enrolment',
        text='Percentage',
        barmode='group'
    )
    fig12.update_layout(
        autosize=False,
        width=1200,  # Increased the width
        height=800,  # Increased the height
        legend_title_text='Departments'
    )
    fig12.update_traces(texttemplate='%{text}', textposition='outside')

    # Load data from Sheet11
    df11 = pd.read_excel(students_file_path, sheet_name='Sheet11')
    df11.columns = df11.columns.str.strip()

    # Melt the DataFrame for better plotting
    df11_melted = df11.melt(id_vars='Department', value_vars=df11.columns[1:], var_name='Year', value_name='No. of Postgraduate enrolment')

    # Create the bar graph for Sheet11
    fig13 = px.bar(
        df11_melted,
        x='Year',
        y='No. of Postgraduate enrolment',
        color='Department',
        title='International Students Postgraduate Enrolment - Actual Numbers',
        text='No. of Postgraduate enrolment',
        barmode='group'
    )
    fig13.update_layout(
        autosize=False,
        width=1200,  # Increased the width
        height=800,  # Increased the height
        legend_title_text='Departments'
    )
    fig13.update_traces(texttemplate='%{text}', textposition='outside')

    return fig1, fig2, fig3, fig4, fig5, fig6, fig7, fig8, fig9, fig10, fig11, fig12, fig13

# Functions to create charts for student performance data

def create_student_performance_charts():
    # Read the Excel files into DataFrames
    df1 = pd.read_excel(student_performance_file_path, sheet_name='Sheet1')
    df2 = pd.read_excel(student_performance_file_path, sheet_name='Sheet2')
    df3 = pd.read_excel(student_performance_file_path, sheet_name='Sheet3')
    df4 = pd.read_excel(student_performance_file_path, sheet_name='Sheet4')
    df5 = pd.read_excel(student_performance_file_path, sheet_name='Sheet5')
    df6 = pd.read_excel(student_performance_file_path, sheet_name='Sheet6')
    df7 = pd.read_excel(student_performance_file_path, sheet_name='Sheet7')
    df8 = pd.read_excel(student_performance_file_path, sheet_name='Sheet8')
    df9 = pd.read_excel(student_performance_file_path, sheet_name='Sheet9')
    df10 = pd.read_excel(student_performance_file_path, sheet_name='Sheet10')
    df11 = pd.read_excel(student_performance_file_path, sheet_name='Sheet11')
    df12 = pd.read_excel(student_performance_file_path, sheet_name='Sheet12')
    df13 = pd.read_excel(student_performance_file_path, sheet_name='Sheet13')
    df14 = pd.read_excel(student_performance_file_path, sheet_name='Sheet14')
    df15 = pd.read_excel(student_performance_file_path, sheet_name='Sheet15')

    # Ensure Success Rates in Sheet1 are strings and convert them to float
    df1['Success Rates'] = df1['Success Rates'].astype(str).str.rstrip('%').astype(float)

    # Prepare data for Linear Regression from Sheet1
    X1 = df1[['Year']]
    y1 = df1['Success Rates']

    # Impute any missing values in X1 and y1
    imputer_X1 = SimpleImputer(strategy='mean')
    imputer_y1 = SimpleImputer(strategy='mean')
    X1 = imputer_X1.fit_transform(X1)
    y1 = imputer_y1.fit_transform(y1.values.reshape(-1, 1)).ravel()

    # Fit the Linear Regression model for Sheet1
    model1 = LinearRegression()
    model1.fit(X1, y1)

    # Forecasting for the next 5 years for Sheet1
    future_years1 = np.arange(df1['Year'].max() + 1, df1['Year'].max() + 6).reshape(-1, 1)
    predictions1 = model1.predict(future_years1)

    # Append predictions to the DataFrame for Sheet1
    future_df1 = pd.DataFrame({'Year': future_years1.flatten(), 'Success Rates': predictions1})

    # Ensure Success Rates in Sheet2 are strings and convert them to float
    for year in ['2019', '2020', '2021', '2022', '2023']:
        df2[year] = df2[year].astype(str).str.rstrip('%').astype(float)

    # Melt the DataFrame from Sheet2 to have long-form data for easier plotting
    df_melted2 = df2.melt(id_vars=['Department'], value_vars=['2019', '2020', '2021', '2022', '2023'],
                          var_name='Year', value_name='Success Rates')

    # Ensure Success Rates in Sheet3 are strings and convert them to float
    df3['Success Rates of First Time Entering Students'] = df3['Success Rates of First Time Entering Students'].astype(str).str.rstrip('%').astype(float)

    # Prepare data for Linear Regression from Sheet3
    X3 = df3[['Year']]
    y3 = df3['Success Rates of First Time Entering Students']

    # Impute any missing values in X3 and y3
    imputer_X3 = SimpleImputer(strategy='mean')
    imputer_y3 = SimpleImputer(strategy='mean')
    X3 = imputer_X3.fit_transform(X3)
    y3 = imputer_y3.fit_transform(y3.values.reshape(-1, 1)).ravel()

    # Fit the Linear Regression model for Sheet3
    model3 = LinearRegression()
    model3.fit(X3, y3)

    # Forecasting for the next 5 years for Sheet3
    future_years3 = np.arange(df3['Year'].max() + 1, df3['Year'].max() + 6).reshape(-1, 1)
    predictions3 = model3.predict(future_years3)

    # Append predictions to the DataFrame for Sheet3
    future_df3 = pd.DataFrame({'Year': future_years3.flatten(), 'Success Rates': predictions3})

    # Ensure Success Rates in Sheet4 are strings and convert them to float
    df4['Success Rates of African Students'] = df4['Success Rates of African Students'].astype(str).str.rstrip('%').astype(float)

    # Prepare data for Linear Regression from Sheet4
    X4 = df4[['Year']]
    y4 = df4['Success Rates of African Students']

    # Impute any missing values in X4 and y4
    imputer_X4 = SimpleImputer(strategy='mean')
    imputer_y4 = SimpleImputer(strategy='mean')
    X4 = imputer_X4.fit_transform(X4)
    y4 = imputer_y4.fit_transform(y4.values.reshape(-1, 1)).ravel()

    # Fit the Linear Regression model for Sheet4
    model4 = LinearRegression()
    model4.fit(X4, y4)

    # Forecasting for the next 5 years for Sheet4
    future_years4 = np.arange(df4['Year'].max() + 1, df4['Year'].max() + 6).reshape(-1, 1)
    predictions4 = model4.predict(future_years4)

    # Append predictions to the DataFrame for Sheet4
    future_df4 = pd.DataFrame({'Year': future_years4.flatten(), 'Success Rates': predictions4})

    # Ensure Success Rates in Sheet5 are strings and convert them to float
    df5['FACULTY'] = df5['FACULTY'].astype(str).str.rstrip('%').astype(float)

    # Ensure Success Rates in Sheet6 are strings and convert them to float
    for year in ['2014', '2015', '2016', '2017', '2018', '2019', '2020', '2021', '2022', ]:
        df6[year] = df6[year].astype(str).str.rstrip('%').astype(float)
    df6['Difference: 2014 vs 2022'] = df6['Difference: 2014 vs 2022'].astype(str).str.rstrip('%').astype(float)

    # Ensure Success Rates in Sheet7 are strings and convert them to float
    for year in ['2015', '2016', '2017', '2018', '2019', '2020', '2021', '2022', ]:
        df7[year] = df7[year].astype(str).str.rstrip('%').astype(float)

    # Ensure Success Rates in Sheet8 are strings and convert them to float
    for year in ['2014', '2015', '2016', '2017', '2018', '2019', '2020', '2021', '2022', ]:
        df8[year] = df8[year].astype(str).str.rstrip('%').astype(float)
    df8['Difference: 2014 vs 2022'] = df8['Difference: 2014 vs 2022'].astype(str).str.rstrip('%').astype(float)

    # Ensure Success Rates in Sheet9 are strings and convert them to float
    for year in ['2014', '2015', '2016', '2017', '2018', '2019', '2020']:
        df9[year] = df9[year].astype(str).str.rstrip('%').astype(float)

    # Ensure Success Rates in Sheet10 are strings and convert them to float
    df10['Dropout'] = df10['Dropout'].astype(str).str.rstrip('%').astype(float)
    df10['Throughput'] = df10['Throughput'].astype(str).str.rstrip('%').astype(float)
    df10['Still in Progress'] = df10['Still in Progress'].astype(str).str.rstrip('%').astype(float)

    # Ensure Success Rates in Sheet11 are strings and convert them to float
    for year in ['2014', '2015', '2016', '2019', '2020', '2021', '2022', ]:
        df11[year] = df11[year].astype(str).str.rstrip('%').astype(float)

    # Ensure Success Rates in Sheet12 are strings and convert them to float
    df12['Faculty'] = df12['Faculty'].astype(str).str.rstrip('%').astype(float)

    # Filter out the row containing "Difference: 2014 vs. 2021"
    df12_filtered = df12[df12['Year'] != 'Difference: 2014 vs. 2021']
    df12_filtered['Year'] = df12_filtered['Year'].astype(int)

    # Prepare data for Linear Regression from Sheet12
    X12 = df12_filtered[['Year']]
    y12 = df12_filtered['Faculty']

    # Impute any missing values in X12 and y12
    imputer_X12 = SimpleImputer(strategy='mean')
    imputer_y12 = SimpleImputer(strategy='mean')
    X12 = imputer_X12.fit_transform(X12)
    y12 = imputer_y12.fit_transform(y12.values.reshape(-1, 1)).ravel()

    # Fit the Linear Regression model for Sheet12
    model12 = LinearRegression()
    model12.fit(X12, y12)

    # Forecasting for the next 5 years for Sheet12
    future_years12 = np.arange(df12_filtered['Year'].max() + 1, df12_filtered['Year'].max() + 6).reshape(-1, 1)
    predictions12 = model12.predict(future_years12)

    # Append predictions to the DataFrame for Sheet12
    future_df12 = pd.DataFrame({'Year': future_years12.flatten(), 'Faculty': predictions12})

    # Ensure Success Rates in Sheet13 are strings and convert them to float
    for year in ['2014', '2015', '2016', '2017', '2018', '2019', '2020', '2021', '2022', ]:
        df13[year] = df13[year].astype(str).str.rstrip('%').astype(float)
    df13['Difference: 2014 vs 2023'] = df13['Difference: 2014 vs 2023'].astype(str).str.rstrip('%').astype(float)

    # Ensure Success Rates in Sheet14 are strings and convert them to float
    df14['Faculty'] = df14['Faculty'].astype(str).str.rstrip('%').astype(float)

    # Prepare data for Linear Regression from Sheet14
    X14 = df14[['Year']]
    y14 = df14['Faculty']

    # Impute any missing values in X14 and y14
    imputer_X14 = SimpleImputer(strategy='mean')
    imputer_y14 = SimpleImputer(strategy='mean')
    X14 = imputer_X14.fit_transform(X14)
    y14 = imputer_y14.fit_transform(y14.values.reshape(-1, 1)).ravel()

    # Fit the Linear Regression model for Sheet14
    model14 = LinearRegression()
    model14.fit(X14, y14)

    # Forecasting for the next 5 years for Sheet14
    future_years14 = np.arange(df14['Year'].max() + 1, df14['Year'].max() + 6).reshape(-1, 1)
    predictions14 = model14.predict(future_years14)

    # Append predictions to the DataFrame for Sheet14
    future_df14 = pd.DataFrame({'Year': future_years14.flatten(), 'Faculty': predictions14})

    # Ensure Success Rates in Sheet15 are strings and convert them to float
    for year in ['2014', '2015', '2016', '2017', '2018', '2019', '2020', '2021', '2022', '2023']:
        df15[year] = df15[year].astype(str).str.rstrip('%').astype(float)

    # Find the year with the highest overall performance in Sheet2
    highest_year = df_melted2.loc[df_melted2['Success Rates'].idxmax()]['Year']

    # Create the graphs for each sheet
    graphs = [
        go.Figure(
            data=[
                go.Scatter(
                    x=df1['Year'],
                    y=df1['Success Rates'],
                    mode='lines+markers',
                    name='Actual Success Rate',
                    line=dict(color='blue')
                ),
                go.Scatter(
                    x=future_df1['Year'],
                    y=future_df1['Success Rates'],
                    mode='lines+markers',
                    name='Forecasted Success Rate',
                    line=dict(color='red', dash='dash')
                )
            ],
            layout=go.Layout(
                title='FAS Overall Student Success Rate (Sheet1)',
                xaxis={'title': 'Year'},
                yaxis={'title': 'Success Rate (%)'}
            )
        ),
        go.Figure(
            data=[
                go.Bar(
                    x=df_melted2[df_melted2['Department'] == dept]['Year'],
                    y=df_melted2[df_melted2['Department'] == dept]['Success Rates'],
                    name=dept
                ) for dept in df2['Department']
            ],
            layout=go.Layout(
                title='Department Success Rates by Year (Sheet2)',
                xaxis={'title': 'Year'},
                yaxis={'title': 'Success Rate (%)'},
                barmode='group'
            )
        ),
        go.Figure(
            data=[
                go.Bar(
                    x=df_melted2[df_melted2['Department'] == dept]['Year'],
                    y=df_melted2[df_melted2['Department'] == dept]['Success Rates'],
                    name=dept
                ) for dept in df2['Department']
            ],
            layout=go.Layout(
                title='Department Success Rates by Year with Highlight (Sheet2)',
                xaxis={'title': 'Year'},
                yaxis={'title': 'Success Rate (%)'},
                barmode='group',
                annotations=[
                    dict(
                        x=highest_year,
                        y=df_melted2['Success Rates'].max(),
                        xref='x',
                        yref='y',
                        text='Highest Performance Year',
                        showarrow=True,
                        arrowhead=7,
                        ax=0,
                        ay=-40
                    )
                ]
            )
        ),
        go.Figure(
            data=[
                go.Scatter(
                    x=df3['Year'],
                    y=df3['Success Rates of First Time Entering Students'],
                    mode='lines+markers',
                    name='Actual Success Rate',
                    line=dict(color='green')
                ),
                go.Scatter(
                    x=future_df3['Year'],
                    y=future_df3['Success Rates'],
                    mode='lines+markers',
                    name='Forecasted Success Rate',
                    line=dict(color='orange', dash='dash')
                )
            ],
            layout=go.Layout(
                title='Success Rates of First Time Entering Students (Sheet3)',
                xaxis={'title': 'Year'},
                yaxis={'title': 'Success Rate (%)'}
            )
        ),
        go.Figure(
            data=[
                go.Scatter(
                    x=df4['Year'],
                    y=df4['Success Rates of African Students'],
                    mode='lines+markers',
                    name='Actual Success Rate',
                    line=dict(color='purple')
                ),
                go.Scatter(
                    x=future_df4['Year'],
                    y=future_df4['Success Rates'],
                    mode='lines+markers',
                    name='Forecasted Success Rate',
                    line=dict(color='brown', dash='dash')
                )
            ],
            layout=go.Layout(
                title='Success Rates of African Students (Sheet4)',
                xaxis={'title': 'Year'},
                yaxis={'title': 'Success Rate (%)'}
            )
        ),
        go.Figure(
            data=[
                go.Scatter(
                    x=df5['Year'],
                    y=df5['FACULTY'],
                    mode='lines+markers',
                    name='FACULTY Success Rate',
                    line=dict(color='cyan')
                )
            ],
            layout=go.Layout(
                title='Faculty Student Throughput - Undergraduate',
                xaxis={'title': 'Year'},
                yaxis={'title': 'Success Rate (%)'}
            )
        ),
        go.Figure(
            data=[
                go.Bar(
                    x=[year for year in range(2014, 2022)],
                    y=df6.loc[df6['Department'] == dept, [str(year) for year in range(2014, 2022)]].values.flatten(),
                    name=dept
                ) for dept in df6['Department']
            ],
            layout=go.Layout(
                title='Department Success Rates by Year (Sheet6)',
                xaxis={'title': 'Year'},
                yaxis={'title': 'Success Rate (%)'},
                barmode='group'
            )
        ),
        go.Figure(
            data=[
                go.Bar(
                    x=df6['Department'],
                    y=df6['Difference: 2014 vs 2022'],
                    name='Difference 2014 vs 2022',
                    marker=dict(color='blue')
                )
            ],
            layout=go.Layout(
                title='Difference in Success Rates 2014 vs 2022 (Sheet6)',
                xaxis={'title': 'Department'},
                yaxis={'title': 'Difference (%)'}
            )
        ),
        go.Figure(
            data=[
                go.Bar(
                    x=[str(year) for year in range(2015, 2022)],
                    y=df7.loc[df7['Department'] == dept, [str(year) for year in range(2015, 2022)]].values.flatten(),
                    name=dept
                ) for dept in df7[df7['Department'].str.contains('Masters')]['Department']
            ],
            layout=go.Layout(
                title='Postgraduate Throughput - Masters',
                xaxis={'title': 'Year'},
                yaxis={'title': 'Success Rate (%)'},
                barmode='group'
            )
        ),
        go.Figure(
            data=[
                go.Bar(
                    x=[str(year) for year in range(2015, 2022)],
                    y=df7.loc[df7['Department'] == dept, [str(year) for year in range(2015, 2022)]].values.flatten(),
                    name=dept
                ) for dept in df7[df7['Department'].str.contains('PhD')]['Department']
            ],
            layout=go.Layout(
                title='Postgraduate Throughput - PhD',
                xaxis={'title': 'Year'},
                yaxis={'title': 'Success Rate (%)'},
                barmode='group'
            )
        ),
        go.Figure(
            data=[
                go.Bar(
                    x=[str(year) for year in range(2014, 2022)],
                    y=df8.loc[df8['Department'] == dept, [str(year) for year in range(2014, 2022)]].values.flatten(),
                    name=dept
                ) for dept in df8['Department']
            ],
            layout=go.Layout(
                title='Student Dropout Rates - Undergraduate',
                xaxis={'title': 'Year'},
                yaxis={'title': 'Dropout Rate (%)'},
                barmode='group'
            )
        ),
        go.Figure(
            data=[
                go.Bar(
                    x=[str(year) for year in range(2014, 2021)],
                    y=df9.loc[df9['Department'] == dept, [str(year) for year in range(2014, 2021)]].values.flatten(),
                    name=dept
                ) for dept in df9['Department']
            ],
            layout=go.Layout(
                title='Dropout Rate in The First Year',
                xaxis={'title': 'Year'},
                yaxis={'title': 'Dropout Rate (%)'},
                barmode='group'
            )
        ),
        go.Figure(
            data=[
                go.Bar(
                    x=df10['Department'],
                    y=df10['Dropout'],
                    name='Dropout'
                ),
                go.Bar(
                    x=df10['Department'],
                    y=df10['Throughput'],
                    name='Throughput'
                ),
                go.Bar(
                    x=df10['Department'],
                    y=df10['Still in Progress'],
                    name='Still in Progress'
                )
            ],
            layout=go.Layout(
                title='Dropout, Throughput, and Still in Progress (Sheet10)',
                xaxis={'title': 'Department'},
                yaxis={'title': 'Percentage (%)'},
                barmode='group'
            )
        ),
        go.Figure(
            data=[
                go.Bar(
                    x=[str(year) for year in range(2014, 2024) if str(year) in df11.columns],
                    y=df11.loc[df11['Department'] == dept, [str(year) for year in range(2014, 2024) if str(year) in df11.columns]].values.flatten(),
                    name=dept
                ) for dept in df11[df11['Department'].str.contains('Masters')]['Department']
            ],
            layout=go.Layout(
                title='Postgraduate Dropout - Masters',
                xaxis={'title': 'Year'},
                yaxis={'title': 'Dropout Rate (%)'},
                barmode='group'
            )
        ),
        go.Figure(
            data=[
                go.Bar(
                    x=[str(year) for year in range(2014, 2024) if str(year) in df11.columns],
                    y=df11.loc[df11['Department'] == dept, [str(year) for year in range(2014, 2024) if str(year) in df11.columns]].values.flatten(),
                    name=dept
                ) for dept in df11[df11['Department'].str.contains('PhD')]['Department']
            ],
            layout=go.Layout(
                title='Postgraduate Dropout - PhD',
                xaxis={'title': 'Year'},
                yaxis={'title': 'Dropout Rate (%)'},
                barmode='group'
            )
        ),
        go.Figure(
            data=[
                go.Scatter(
                    x=df12_filtered['Year'],
                    y=df12_filtered['Faculty'],
                    mode='lines+markers',
                    name='Actual Graduation Rate',
                    line=dict(color='blue')
                ),
                go.Scatter(
                    x=future_df12['Year'],
                    y=future_df12['Faculty'],
                    mode='lines+markers',
                    name='Forecasted Graduation Rate',
                    line=dict(color='red', dash='dash')
                )
            ],
            layout=go.Layout(
                title='FAS Graduation Rates',
                xaxis={'title': 'Year'},
                yaxis={'title': 'Graduation Rate (%)'},
                annotations=[
                    dict(
                        x=2021,
                        y=df12.loc[df12['Year'] == '2021', 'Faculty'].values[0],
                        xref='x',
                        yref='y',
                        text='Difference: 2014 vs 2021 = 13%',
                        showarrow=True,
                        arrowhead=7,
                        ax=0,
                        ay=-40
                    )
                ]
            )
        ),
        go.Figure(
            data=[
                go.Bar(
                    x=[year for year in range(2014, 2024)],
                    y=df13.loc[df13['Department'] == dept, [str(year) for year in range(2014, 2024)]].values.flatten(),
                    name=dept
                ) for dept in df13['Department']
            ],
            layout=go.Layout(
                title='Graduation Rates By Programme',
                xaxis={'title': 'Year'},
                yaxis={'title': 'Graduation Rate (%)'},
                barmode='group',
                annotations=[
                    dict(
                        x=2021,
                        y=df13[[str(year) for year in range(2014, 2024)]].max().max(),
                        xref='x',
                        yref='y',
                        text='Year with Most Graduates',
                        showarrow=True,
                        arrowhead=7,
                        ax=0,
                        ay=-40
                    )
                ]
            )
        ),
        go.Figure(
            data=[
                go.Bar(
                    x=df13['Department'],
                    y=df13['Difference: 2014 vs 2023'],
                    name='Difference 2014 vs 2023',
                    marker=dict(color='blue')
                )
            ],
            layout=go.Layout(
                title='Difference in Graduation Rates 2014 vs 2022 (Sheet13)',
                xaxis={'title': 'Department'},
                yaxis={'title': 'Difference (%)'}
            )
        ),
        go.Figure(
            data=[
                go.Scatter(
                    x=df14['Year'],
                    y=df14['Faculty'],
                    mode='lines+markers',
                    name='Actual Graduation Rate',
                    line=dict(color='blue')
                ),
                go.Scatter(
                    x=future_df14['Year'],
                    y=future_df14['Faculty'],
                    mode='lines+markers',
                    name='Forecasted Graduation Rate',
                    line=dict(color='red', dash='dash')
                )
            ],
            layout=go.Layout(
                title='Postgraduate Graduation Rate',
                xaxis={'title': 'Year'},
                yaxis={'title': 'Graduation Rate (%)'}
            )
        ),
        go.Figure(
            data=[
                go.Bar(
                    x=[year for year in range(2014, 2024)],
                    y=df15.loc[df15['Department'] == dept, [str(year) for year in range(2014, 2024)]].values.flatten(),
                    name=dept
                ) for dept in df15['Department']
            ],
            layout=go.Layout(
                title='Pass Rates by Department (Sheet15)',
                xaxis={'title': 'Year'},
                yaxis={'title': 'Pass Rate (%)'},
                barmode='group'
            )
        )
    ]

    return graphs


def create_dut_charts(selected_graph):
     
  app.layout = html.Div([
    html.H1(""),
    
    dcc.Dropdown(
        id='graph-selector',
        options=[
            {'label': 'Postgraduate Enrolment (2020-2023)', 'value': 'graph1'},
            {'label': 'FAS Postgraduate Enrolment (2020-2023)', 'value': 'graph2'},
            {'label': '2023 Student Enrolment by Level', 'value': 'graph3'},
            {'label': 'Postgraduate Graduation Rate (2015-2023)', 'value': 'graph4'},
            {'label': 'Postgraduate Enrolment 2024 (Image)', 'value': 'image1'},
            {'label': 'Current Postdoctoral Fellows (Image)', 'value': 'image2'},
            {'label': 'Emeritus/Honorary/Adjunct Professors (Image)', 'value': 'image3'},
            {'label': 'Departmental Research Outputs 2023 (Image)', 'value': 'image4'}
        ],
        value='graph1'
    ),
    
    dcc.Graph(id='graph-output', style={'display': 'block'}),
    html.Img(id='image-output', style={'display': 'none', 'width': '80%', 'height': 'auto'}),
    html.Div(id='image-title', style={'text-align': 'center', 'font-size': '20px', 'margin-top': '10px'})
])

# Callback function to update the graph or image based on the dropdown selection
@app.callback(
    [Output('graph-output', 'figure'), Output('graph-output', 'style'),
     Output('image-output', 'src'), Output('image-output', 'style'),
     Output('image-title', 'children')],
    Input('graph-selector', 'value')
)
def update_output(selected_graph):
    if selected_graph == 'graph1':
        # Filter data for Sheet1
        df_filtered = df_sheet1[['Postgraduate Enrolment', '2020', '2021', '2022', '2023']].copy()
        df_filtered.set_index('Postgraduate Enrolment', inplace=True)
        
        # Create the bar chart for graph1
        fig = go.Figure()
        for col in df_filtered.columns:
            fig.add_trace(go.Bar(
                x=df_filtered.index,
                y=df_filtered[col],
                name=col,
                text=df_filtered[col],  # Display actual values
                textposition='auto'
            ))
        fig.update_layout(
            title='Postgraduate Enrolment - Actual Student Numbers (2020-2023)',
            xaxis_title='Subjects',
            yaxis_title='Number of Students',
            barmode='group'
        )
        return fig, {'display': 'block'}, None, {'display': 'none'}, ''
    
    elif selected_graph == 'graph2':
        # Filter data for Sheet2
        df_filtered = df_sheet2[['FAS Postgraduate Enrolment', '2020', '2021', '2022', '2023']].copy()
        df_filtered.set_index('FAS Postgraduate Enrolment', inplace=True)
        df_filtered *= 100  # Convert values to percentage
        
        # Create the bar chart for graph2
        fig = go.Figure()
        for col in df_filtered.columns:
            fig.add_trace(go.Bar(
                x=df_filtered.index,
                y=df_filtered[col],
                name=col,
                text=[f'{val:.0f}%' for val in df_filtered[col]],  # Display percentages
                textposition='auto'
            ))
        fig.update_layout(
            title='FAS Postgraduate Enrolment (2020-2023)',
            xaxis_title='Category',
            yaxis_title='Enrolment (%)',
            barmode='group'
        )
        return fig, {'display': 'block'}, None, {'display': 'none'}, ''
    
    elif selected_graph == 'graph3':
        # Filter data for Sheet3
        df_filtered = df_sheet3[['2023 Student Enrolment by Level', 'UG (NQF 5-7)', 'PG upto Masters (NQF8)', 'PG (NQF9-10)']].copy()
        df_filtered.set_index('2023 Student Enrolment by Level', inplace=True)
        
        # Create the bar chart for graph3
        fig = go.Figure()
        for col in df_filtered.columns:
            fig.add_trace(go.Bar(
                x=df_filtered.index,
                y=df_filtered[col],
                name=col,
                text=df_filtered[col],  # Display actual values
                textposition='auto'
            ))
        fig.update_layout(
            title='2023 Student Enrolment by Level',
            xaxis_title='Programs',
            yaxis_title='Number of Students',
            barmode='group'
        )
        return fig, {'display': 'block'}, None, {'display': 'none'}, ''
    
    elif selected_graph == 'graph4':
        # Strip any leading/trailing spaces from column names and ensure integer type for graduation rate
        df_sheet4.columns = df_sheet4.columns.str.strip()
        df_sheet4['Postgraduate Graduation Rate'] = df_sheet4['Postgraduate Graduation Rate'].astype(int)
        
        # Create the line graph for graph4
        fig = go.Figure(data=go.Scatter(
            x=df_sheet4['Postgraduate Graduation Rate'],
            y=df_sheet4['Faculty'],
            mode='lines+markers',
            text=[f'{val}%' for val in df_sheet4['Faculty']],  # Display percentage values
            textposition='bottom center',
            line=dict(color='blue')
        ))
        fig.update_layout(
            title='Postgraduate Graduation Rate (2015-2023)',
            xaxis_title='Year',
            yaxis_title='Graduation Rate (%)',
            xaxis=dict(tickmode='linear')  # Ensure all years are displayed
        )
        return fig, {'display': 'block'}, None, {'display': 'none'}, ''
    
    elif selected_graph == 'image1':
        # Display first image (Postgraduate Enrolment 2024)
        return {}, {'display': 'none'}, '/content/1.png', {'display': 'block'}, 'Postgraduate Enrolment 2024'
    
    elif selected_graph == 'image2':
        # Display second image (Current Postdoctoral Fellows)
        return {}, {'display': 'none'}, '/content/2.png', {'display': 'block'}, 'Current Postdoctoral Fellows'
    
    elif selected_graph == 'image3':
        # Display third image (Emeritus/Honorary/Adjunct Professors)
        return {}, {'display': 'none'}, '/content/3.png', {'display': 'block'}, 'Emeritus/Honorary/Adjunct Professors'
    
    elif selected_graph == 'image4':
        # Display fourth image (Departmental Research Outputs 2023)
        return {}, {'display': 'none'}, '/content/4.png', {'display': 'block'}, 'Departmental Research Outputs 2023'


# Dropdown options for the staff and student pages
staff_dropdown_options = [
    {'label': 'Faculty Data (Sheet1)', 'value': 'Sheet1'},
    {'label': 'Department Success Rates (Sheet2)', 'value': 'Sheet2'},
    {'label': 'Academic Staff with PhD (Sheet3)', 'value': 'Sheet3'},
    {'label': 'Forecast for the Next 5 Years - Faculty Data (Sheet1)', 'value': 'Sheet1_forecast'},
    {'label': 'Forecast for the Next 5 Years - Department Success Rates (Sheet2)', 'value': 'Sheet2_forecast'},
    {'label': 'Forecast for the Next 5 Years - Academic Staff with PhD (Sheet3)', 'value': 'Sheet3_forecast'},
    {'label': 'Department Success Rates Line (Sheet4)', 'value': 'Sheet4'},
    {'label': 'Percentage Difference (2022 vs. 2014)', 'value': 'Sheet5_diff'},
    {'label': 'Percentage of Full-Time Permanent Academic Staff with PhD (2014-2022)', 'value': 'Sheet5'}
]


students_dropdown_options = [
    {'label': 'Headcount Enrolment: Planned vs Achieved (2014-2022)', 'value': 'fig1'},
    {'label': 'Linear Regression Forecast for the Next 5 Years', 'value': 'fig2'},
    {'label': 'Enrolment by Level', 'value': 'fig3'},
    {'label': 'Number of Students by Department (2014 vs 2022)', 'value': 'fig4'},
    {'label': '% African Students (2014-2022)', 'value': 'fig5'},
    {'label': '% Female Students (2014-2022)', 'value': 'fig6'},
    {'label': 'Postgraduate Enrolment (2014-2022)', 'value': 'fig7'},
    {'label': 'Postgraduate (M+D) Enrolment by Department', 'value': 'fig8'},
    {'label': 'Difference 2014 vs 2022 by Department', 'value': 'fig9'},
    {'label': 'Postgraduate Enrolment - Actual Student Numbers', 'value': 'fig10'},
    {'label': 'International Student Postgraduate Enrolment', 'value': 'fig11'},
    {'label': 'International Students Postgraduate Enrolment - Actual Numbers', 'value': 'fig12'}
]

# Dropdown options for the performance page
performance_dropdown_options = [
    {'label': 'FAS Overall Student Success Rate', 'value': 'FAS Overall Student Success Rate (Sheet1)'},
    {'label': 'Department Success Rates by Year', 'value': 'Department Success Rates by Year (Sheet2)'},
    {'label': 'Success Rates of First Time Entering Students', 'value': 'Success Rates of First Time Entering Students (Sheet3)'},
    {'label': 'Success Rates of African Students', 'value': 'Success Rates of African Students (Sheet4)'},
    {'label': 'Faculty Student Throughput - Undergraduate', 'value': 'Faculty Student Throughput - Undergraduate (Sheet5)'},
    {'label': 'Department Success Rates by Year', 'value': 'Department Success Rates by Year (Sheet6)'},
    {'label': 'Difference in Success Rates 2014 vs 2022', 'value': 'Difference in Success Rates 2014 vs 2022 (Sheet6)'},
    {'label': 'Postgraduate Throughput - Masters', 'value': 'Postgraduate Throughput - Masters (Sheet7)'},
    {'label': 'Postgraduate Throughput - PhD', 'value': 'Postgraduate Throughput - PhD (Sheet7)'},
    {'label': 'Student Dropout Rates - Undergraduate', 'value': 'Student Dropout Rates - Undergraduate (Sheet8)'},
    {'label': 'Dropout Rate in The First Year', 'value': 'Dropout Rate in The First Year (Sheet9)'},
    {'label': 'Dropout, Throughput, and Still in Progress', 'value': 'Dropout, Throughput, and Still in Progress (Sheet10)'},
    {'label': 'Postgraduate Dropout - Masters', 'value': 'Postgraduate Dropout - Masters (Sheet11)'},
    {'label': 'Postgraduate Dropout - PhD', 'value': 'Postgraduate Dropout - PhD (Sheet11)'},
    {'label': 'FAS Graduation Rates', 'value': 'FAS Graduation Rates (Sheet12)'},
    {'label': 'Graduation Rates By Programme', 'value': 'Graduation Rates By Programme (Sheet13)'},
    {'label': 'Difference in Graduation Rates 2014 vs 2022', 'value': 'Difference in Graduation Rates 2014 vs 2022 (Sheet13)'},
    {'label': 'Postgraduate Graduation Rate', 'value': 'Postgraduate Graduation Rate (Sheet14)'},
    {'label': 'Pass Rates by Department', 'value': 'Pass Rates by Department (Sheet15)'}
]

# Staff page layout
bar_figures, line_figures, forecast_figures, fig_diff, fig_sheet5 = create_staff_charts()
staff_layout = html.Div(style={'textAlign': 'center'}, children=[
    html.H1("Staff Preliminary Analysis"),
    html.Img(src='/assets/my_image.png', style={'position': 'absolute', 'top': '10px', 'left': '10px', 'width': '150px', 'height': 'auto'}),
    dcc.Dropdown(
        id='staff-dropdown',
        options=staff_dropdown_options,
        value='Sheet1',
        style={'width': '50%', 'margin': 'auto'}
    ),
    dcc.Graph(id='staff-graph')
])

# Students page layout
fig1, fig2, fig3, fig4, fig5, fig6, fig7, fig8, fig9, fig10, fig11, fig12, fig13 = create_students_charts()
students_layout = html.Div(style={'textAlign': 'center'}, children=[
    html.H1("Students Preliminary Analysis"),
    dcc.Dropdown(
        id='students-dropdown',
        options=students_dropdown_options,
        value='fig1',
        style={'width': '50%', 'margin': 'auto'}
    ),
    dcc.Graph(id='students-graph')
])

# Performance page layout
performance_figures = create_student_performance_charts()
performance_layout = html.Div(style={'textAlign': 'center'}, children=[
    html.H1("Student Performance Indicators"),
    dcc.Dropdown(
        id='performance-dropdown',
        options=performance_dropdown_options,
        value='FAS Overall Student Success Rate (Sheet1)',
        style={'width': '50%', 'margin': 'auto'}
    ),
    dcc.Graph(id='performance-graph')
])


# DUT Research layout
dut_layout = html.Div(style={'textAlign': 'center'}, children=[
    html.H1("DUT Research Dashboard"),
    dcc.Dropdown(id='dut-dropdown', options=dut_dropdown_options, value='graph1', style={'width': '50%', 'margin': 'auto'}),
    dcc.Graph(id='dut-graph')
])

# Chatbot layout
chatbot_layout = html.Div(style={'textAlign': 'center'}, children=[
    html.H1("Chat with Us"),
    dcc.Textarea(
        id='chat-input',
        style={'width': '50%', 'height': '100px', 'margin': 'auto'}
    ),
    html.Button('Send', id='send-button', n_clicks=0),
    html.Div(id='chat-output')
])

# Main layout with navigation
app.layout = html.Div([
    dcc.Location(id='url', refresh=False),
    html.Div([
        dcc.Link('Staff Preliminary Analysis', href='/'),
        html.Span(' | '),
        dcc.Link('Students Preliminary Analysis', href='/students'),
        html.Span(' | '),
        dcc.Link('Student Performance Indicators', href='/performance'),
        html.Span(' | '),
        dcc.Link('DUT Research Dashboard', href='/dut')
    ], style={'textAlign': 'center', 'margin': '20px'}),
    html.Div(id='page-content')
])

# Callbacks for page navigation
@app.callback(Output('page-content', 'children'), [Input('url', 'pathname')])
def display_page(pathname):
    if pathname == '/students':
        return students_layout
    elif pathname == '/performance':
        return performance_layout
    elif pathname == '/dut':
        return dut_layout
    else:
        return staff_layout

# Callbacks for updating graphs
@app.callback(Output('dut-graph', 'figure'), [Input('dut-dropdown', 'value')])
def update_dut_graph(selected_graph):
    return create_dut_charts(selected_graph)

# Callback for updating the staff graph
@app.callback(Output('staff-graph', 'figure'),
              [Input('staff-dropdown', 'value')])
def update_staff_graph(selected_value):
    if selected_value in bar_figures:
        return bar_figures[selected_value]
    elif selected_value in forecast_figures:
        return forecast_figures[selected_value]
    elif selected_value == 'Sheet5_diff':
        return fig_diff
    elif selected_value == 'Sheet5':
        return fig_sheet5
    else:
        return line_figures[selected_value]

 # Callback for updating the students graph
@app.callback(Output('students-graph', 'figure'),
              [Input('students-dropdown', 'value')])
def update_students_graph(selected_value):
    figures = {
        'fig1': fig1,
        'fig2': fig2,
        'fig3': fig3,
        'fig4': fig4,
        'fig5': fig5,
        'fig6': fig6,
        'fig7': fig7,
        'fig8': fig8,
        'fig9': fig9,
        'fig10': fig10,
        'fig11': fig11,
        'fig12': fig12,
        'fig13': fig13
    }
    return figures[selected_value]

def update_performance_graph(selected_value):
    performance_figures_dict = {
        'FAS Overall Student Success Rate (Sheet1)': performance_figures[0],
        'Department Success Rates by Year (Sheet2)': performance_figures[1],
        'Success Rates of First Time Entering Students (Sheet3)': performance_figures[3],
        'Success Rates of African Students (Sheet4)': performance_figures[4],
        'Faculty Student Throughput - Undergraduate (Sheet5)': performance_figures[5],
        'Department Success Rates by Year (Sheet6)': performance_figures[6],
        'Difference in Success Rates 2014 vs 2022 (Sheet6)': performance_figures[7],
        'Postgraduate Throughput - Masters (Sheet7)': performance_figures[8],
        'Postgraduate Throughput - PhD (Sheet7)': performance_figures[9],
        'Student Dropout Rates - Undergraduate (Sheet8)': performance_figures[10],
        'Dropout Rate in The First Year (Sheet9)': performance_figures[11],
        'Dropout, Throughput, and Still in Progress (Sheet10)': performance_figures[12],
        'Postgraduate Dropout - Masters (Sheet11)': performance_figures[13],
        'Postgraduate Dropout - PhD (Sheet11)': performance_figures[14],
        'FAS Graduation Rates (Sheet12)': performance_figures[15],
        'Graduation Rates By Programme (Sheet13)': performance_figures[16],
        'Difference in Graduation Rates 2014 vs 2022 (Sheet13)': performance_figures[17],
        'Postgraduate Graduation Rate (Sheet14)': performance_figures[18],
        'Pass Rates by Department (Sheet15)': performance_figures[19]
    }
    return performance_figures_dict[selected_value]

# Run the Dash app
if __name__ == '__main__':
    app.run_server(debug=True, port=8050)
